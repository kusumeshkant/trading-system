import io
from dataclasses import asdict
from typing import List, Dict, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.analytics.reports.trade_report import TradeReportData


_HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_WIN_FILL = PatternFill("solid", fgColor="D4EDDA")
_LOSS_FILL = PatternFill("solid", fgColor="F8D7DA")
_ALT_FILL = PatternFill("solid", fgColor="F0F4FF")
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_sheet(ws, header_row: int = 1):
    for cell in ws[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=1):
        fill = _ALT_FILL if i % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 35)

    ws.row_dimensions[header_row].height = 20
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def export_trades_excel(
    reports: List[TradeReportData],
    daily_summaries: Optional[List[Dict]] = None,
    monthly_summaries: Optional[List[Dict]] = None,
) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if reports:
            df = pd.DataFrame([asdict(r) for r in reports])
            df.to_excel(writer, sheet_name="Trade Records", index=False)
            ws = writer.sheets["Trade Records"]
            _style_sheet(ws)

            # Colour P&L rows
            pnl_col = df.columns.get_loc("net_pnl") + 1
            for row_idx, val in enumerate(df["net_pnl"], start=2):
                fill = _WIN_FILL if (val or 0) > 0 else _LOSS_FILL
                ws.cell(row=row_idx, column=pnl_col).fill = fill

        if daily_summaries:
            df_d = pd.DataFrame(daily_summaries)
            df_d.to_excel(writer, sheet_name="Daily Summary", index=False)
            _style_sheet(writer.sheets["Daily Summary"])

        if monthly_summaries:
            df_m = pd.DataFrame(monthly_summaries)
            df_m.to_excel(writer, sheet_name="Monthly Summary", index=False)
            _style_sheet(writer.sheets["Monthly Summary"])

    output.seek(0)
    return output.read()
